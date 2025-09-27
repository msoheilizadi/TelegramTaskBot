import openpyxl
import os
from PIL import Image as PILImage
import subprocess
import fitz
import argparse
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__)) 

def to_persian_number(num):
    s = str(int(round(num)))
    return s.translate(str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹"))


def create_payment_plan_excel(unit_number, discount, payment_method):
    discount = float(discount)
    if discount <= 1:
        discount_pct = discount * 100
    else:
        discount_pct = discount
    discount_frac = discount_pct / 100.0

    # مسیر قالب بسته به نوع پرداخت
    if str(payment_method).strip() == "1":
        template_file = os.path.join(BASE_DIR, "payment_plan_1.xlsx")
    elif str(payment_method).strip() in ("0.5", ".5"):
        template_file = os.path.join(BASE_DIR, "payment_plan_0.5.xlsx")
    else:
        raise ValueError(f"Unknown payment method: {payment_method}")

    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Template not found: {template_file}")

    # مسیر فایل واحدها
    units_file = os.path.join(BASE_DIR, 'units.xlsx')
    if not os.path.exists(units_file):
        raise FileNotFoundError(f"Units file not found: {units_file}")

    units_wb = openpyxl.load_workbook(units_file, data_only=True)
    units_ws = units_wb.active
    payment_wb = openpyxl.load_workbook(template_file)
    payment_ws = payment_wb.active

    unit_data = None
    for row in units_ws.iter_rows(min_row=2):
        if str(row[1].value).strip() == unit_number:
            unit_data = {
                "F": float(row[5].value) if row[5].value is not None else 0,
                "G": row[6].value,
                "M": row[12].value,
                "C": str(row[2].value).strip() if row[2].value is not None else "",
                "I": float(row[8].value) if row[8].value is not None else 0
            }
            break

    if not unit_data:
        raise ValueError(f"Unit number '{unit_number}' not found")

    if unit_data["C"] == "1":
        unit_Type = "یک خوابه"
    elif unit_data["C"] == "2":
        unit_Type = "دو خوابه"
    else:
        unit_Type = "استودیو"

    payment_ws['D9'] = unit_data["I"]

    if not any("D10" in str(rng) for rng in payment_ws.merged_cells.ranges):
        payment_ws.merge_cells('D10:E10')
    payment_ws['D10'] = f"=D9*{discount_frac}"

    if not any("D11" in str(rng) for rng in payment_ws.merged_cells.ranges):
        payment_ws.merge_cells('D11:E11')
    payment_ws['D11'] = "=D9-D10"

    payment_ws['B14'] = unit_number
    payment_ws['C14'] = unit_Type
    payment_ws['D14'] = unit_data["G"]
    payment_ws['E14'] = round(unit_data["F"]) if unit_data["F"] else ""
    payment_ws['F14'] = "=D11"

    if not any("B10" in str(rng) for rng in payment_ws.merged_cells.ranges):
        payment_ws.merge_cells('B10:C10')
    payment_ws['B10'] = f"تخفیف {to_persian_number(discount_pct)} درصدی"

    output_excel = os.path.join(BASE_DIR, f'temp_payment_plan_{unit_number}.xlsx')
    payment_wb.save(output_excel)
    final_aed_price = unit_data["I"] * (1 - discount_frac)
    return output_excel, discount_pct, payment_method, final_aed_price

def excel_to_pdf(excel_path, pdf_path):
    output_dir = os.path.dirname(pdf_path)
    subprocess.run([
        "soffice",
        "--headless",
        "--convert-to", "pdf",
        excel_path,
        "--outdir", output_dir
    ], check=True)


    base_pdf = os.path.splitext(os.path.basename(excel_path))[0] + ".pdf"
    generated_pdf = os.path.join(output_dir, base_pdf)
    if generated_pdf != pdf_path:
        os.rename(generated_pdf, pdf_path)

def add_image_to_pdf_second_page(pdf_path, output_pdf_path, image_path):
    abs_image_path = os.path.abspath(image_path)
    print(f"Checking image path: {abs_image_path}")
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file does not exist: {pdf_path}")
    if not os.path.exists(abs_image_path):
        print(f"Warning: Image not found, skipping image add: {abs_image_path}")
        import shutil
        shutil.copy(pdf_path, output_pdf_path)
        return

    doc = fitz.open(pdf_path)
    if len(doc) < 2:
        doc.new_page(pno=1)
    page = doc[1]

    pw, ph = page.rect.width, page.rect.height
    scale = 1
    pos_x_pct, pos_y_pct = 0.04, 0.455
    width_pct = 0.9

    target_w = pw * width_pct * scale

    with PILImage.open(abs_image_path) as im:
        iw_px, ih_px = im.size
        dpi = im.info.get("dpi", (72,72))[0] if isinstance(im.info.get("dpi", None), tuple) else im.info.get("dpi", 72)
        dpi = dpi or 72
        iw_pt = iw_px * 72.0 / dpi
        ih_pt = ih_px * 72.0 / dpi
        aspect = ih_pt / iw_pt if iw_pt != 0 else (ih_px / iw_px if iw_px != 0 else 1.0)

    target_h = target_w * aspect
    ref_x = pw * pos_x_pct
    ref_y = ph * pos_y_pct

    final_x = ref_x
    final_y = ref_y

    rect = fitz.Rect(final_x, final_y + 5, final_x + target_w + 10, final_y + target_h - 64)

    page.insert_image(rect, filename=abs_image_path, overlay=False)
    doc.save(output_pdf_path)
    doc.close()

def main():
    parser = argparse.ArgumentParser(description="Generate payment plan PDF")
    parser.add_argument('--unit', required=True, help='Unit number')
    parser.add_argument('--discount', required=True, type=float, help='Discount percentage')
    parser.add_argument('--payment', required=True, help='Payment method (1 or 0.5)')
    parser.add_argument('--output', required=True, help='Output PDF path')

    args = parser.parse_args()

    try:
        excel_file, discount_pct, payment_method_val, final_aed_price = create_payment_plan_excel(
            args.unit, args.discount, args.payment
        )

        temp_pdf = f"temp_{args.unit}.pdf"
        excel_to_pdf(excel_file, temp_pdf)

        image_path = os.path.join(BASE_DIR, "images", f"{args.unit}.jpg")

        # Construct final PDF filename with your desired pattern
        final_pdf_name = f"{args.unit}({int(round(discount_pct))}%)({payment_method_val}%).pdf"
        final_pdf_path = os.path.join(BASE_DIR, final_pdf_name)

        add_image_to_pdf_second_page(temp_pdf, final_pdf_path, image_path)

        # cleanup temp files
        if os.path.exists(excel_file):
            os.remove(excel_file)
        if os.path.exists(temp_pdf):
            os.remove(temp_pdf)

        if os.path.exists(final_pdf_path):
            print(final_pdf_path) 
            print(f"AED_PRICE={final_aed_price}")
            sys.exit(0)
        else:
            print(f"Error: PDF file not found at expected path: {final_pdf_path}", file=sys.stderr)
            sys.exit(1)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
    
# def run_generation():
#     unit_number = unit_entry.get().strip()
#     discount = discount_var.get()
#     payment_method = payment_method_var.get()

#     if not unit_number:
#         messagebox.showerror("Input Error", "Please enter a unit number.")
#         return

#     try:
#         excel_file, discount_pct, payment_method_val = create_payment_plan_excel(unit_number, discount, payment_method)
#         pdf_name = f"{unit_number}({int(round(discount_pct))}%)({payment_method_val}%).pdf"
#         pdf_path = os.path.join(BASE_DIR, pdf_name)
#         temp_pdf = os.path.join(BASE_DIR, f"temp_{unit_number}.pdf")
#         excel_to_pdf(excel_file, temp_pdf)
#         image_path = os.path.join(BASE_DIR, "images", f"{unit_number}.jpg")
#         add_image_to_pdf_second_page(temp_pdf, pdf_path, image_path)

#         # cleanup temp files
#         if os.path.exists(excel_file):
#             os.remove(excel_file)
#         if os.path.exists(temp_pdf):
#             os.remove(temp_pdf)

#         messagebox.showinfo("Success", f"Payment plan PDF created:\n{pdf_path}")
#     except Exception as e:
#         messagebox.showerror("Error", str(e))
# # Setup GUI window
# root = tk.Tk()
# root.title("Payment Plan Generator")

# tk.Label(root, text="Unit Number:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
# unit_entry = tk.Entry(root)
# unit_entry.grid(row=0, column=1, padx=5, pady=5)

# tk.Label(root, text="Discount (%):").grid(row=1, column=0, sticky="w", padx=5, pady=5)
# discount_var = tk.StringVar(value="20")
# discount_combo = ttk.Combobox(root, textvariable=discount_var, values=["20", "15", "10", "5", "0"], state="readonly")
# discount_combo.grid(row=1, column=1, padx=5, pady=5)

# tk.Label(root, text="Payment Method (%):").grid(row=2, column=0, sticky="w", padx=5, pady=5)
# payment_method_var = tk.StringVar(value="1")
# payment_method_combo = ttk.Combobox(root, textvariable=payment_method_var, values=["1", "0.5"], state="readonly")
# payment_method_combo.grid(row=2, column=1, padx=5, pady=5)

# generate_btn = tk.Button(root, text="Generate PDF", command=run_generation)
# generate_btn.grid(row=3, column=0, columnspan=2, pady=10)

# root.mainloop()

